"""import time
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
driver=webdriver.Chrome()
driver.get('https://admin-demo.nopcommerce.com/')
driver.maximize_window()



driver.close()"""
import time

"""
driver.find_element(By.ID,'')
driver.find_element(By.XPATH,'')
driver.find_element(By.LINK_TEXT,'')
driver.find_element(By.NAME,'')
driver.find_element(By.CLASS_NAME,'')
driver.find_element(By.CSS_SELECTOR,'')
driver.find_element(By.PARTIAL_LINK_TEXT,'').send_keys()"""
"""import openpyxl
import pyqrcode
import png
from pyqrcode import QRCode


# String which represents the QR code
s = "www.geeksforgeeks.org"

# Generate QR code
url = pyqrcode.create(s)

# Create and save the svg file naming "myqr.svg"
url.svg("myqr.svg", scale = 8)

# Create and save the png file naming "myqr.png"
url.png('myqr.png', scale = 6)
# Give the location of the file
path = "D:\\LoginData1.xlsx"

# workbook object is created
wb_obj = openpyxl.load_workbook(path)


sheet_obj = wb_obj.active

max_col = sheet_obj.max_column

# Will print a particular row value
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = 2, column = i)
    print(cell_obj.value, end = " ")


import segno

video = segno.make('https://docs.google.com/document/d/1Yp5JJ_m13-DImKVHqNmBqFNM0mvANIxM/edit?usp=drivesdk&ouid=106364327883687003401&rtpof=true&sd=true')
video.save('Video.png')"""

"""import numpy as np
arr=np.array([[1,2,3,4,5],[6,7,8,9,10]])
print(arr[1][0])
print(np.max(arr))
print(arr.dtype)
print(arr.shape)
print(arr.size)
print(arr.repeat(5))
for x in arr:
    print(x)"""

"""from selenium import webdriver

from selenium.webdriver.common.by import By

driver = webdriver.Chrome()
driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/pim/viewPersonalDetails/empNumber/7")
driver.find_element(By.NAME,'firstName').send_keys("Admin")
time.sleep(5)
driver.find_element(By.XPATH, '//input[@name="password"]').send_keys("admin123")
driver.find_element(By.XPATH, '//button[text()="Login"]').click()
"""

# import the libs
"""from selenium import webdriver
from time import sleep

# create the initial window
driver = webdriver.Chrome()

# go to the home page
driver.get('https://www.zomato.com')

# storing the current window handle to get back to dashboard
main_page = driver.current_window_handle

# wait for page to load completely
sleep(5)

# click on the sign in tab
driver.find_element_by_xpath('//*[@id ="signin-link"]').click()

sleep(5)

# click to log in using facebook
driver.find_element_by_xpath('//*[@id ="facebook-login-global"]/span').click()

# changing the handles to access login page
for handle in driver.window_handles:
    if handle != main_page:
        login_page = handle

# change the control to signin page
driver.switch_to.window(login_page)

# user input for email and password
print('Enter email id : ', end='')
email = input().strip()
print('Enter password : ', end='')
password = input().strip()

# enter the email
driver.find_element_by_xpath('//*[@id ="email"]').send_keys(email)

# enter the password
driver.find_element_by_xpath('//*[@id ="pass"]').send_keys(password)

# click the login button
driver.find_element_by_xpath('//*[@id ="u_0_0"]').click()

# change control to main page
driver.switch_to.window(main_page)

sleep(10)
# print user name
name = driver.find_element_by_xpath('/html/body/div[4]/div/div[1]/header/div[2]/div/div/div/div/span').text
print('Your user name is : {}'.format(name))

# closing the window
driver.quit()"""

"""from selenium import webdriver
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
#browser exposes an executable file
#Through Selenium test we will invoke the executable file which will then #invoke actual browser
driver = webdriver.Chrome()
# to maximize the browser window
driver.maximize_window()
#get method to launch the URL
driver.get("https://www.tutorialspoint.com/selenium/selenium_automation_practice.htm")
#to refresh the browser
driver.refresh()
#click on submit button
driver.find_element(By.XPATH,"//button[@name='submit']").click()
# alert object creation and switching focus to alert
a = driver.switch_to.alert
# accept the alert
a.accept()
driver.close()"""

"""from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

desired_caps = {
    "app": "C:\\Windows\\System32\\notepad.exe",
}
driver=webdriver.Chrome()

driver = webdriver.Remote(
    command_executor='http://127.0.0.1:4723',
    desired_capabilities=desired_caps)
notepad_window = driver.find_element(By.CLASS_NAME,"Notepad")

# Find the "File" menu
file_menu = notepad_window.find_element(By.NAME,"File")
file_menu.click()

# Click the "Open" option
open_option = notepad_window.find_element(By.NAME,"Open...")
open_option.click()

# Enter the file path
file_path_input = driver.find_element(By.CLASS_NAME,"Edit")
file_path_input.send_keys("C:\\test.txt")

# Click the "Open" button
open_button = driver.find_element(By.NAME,"Open")
open_button.click()
driver.quit()"""

"""from selenium import webdriver
#import Alert class
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.by import By
driver = webdriver.Chrome()
#implicit wait time
driver.implicitly_wait(0.8)
#url launch
driver.get("https://the-internet.herokuapp.com/javascript_alerts")
#identify element
l = driver.find_element(By.XPATH,"//*[text()='Click for JS Prompt']")
l.click()
# instance of Alert class
a = Alert(driver)
# get alert text
print(a.text)
#input text in Alert
a.send_keys('Tutorialspoint')
#dismiss alert
a.dismiss()
l.click()
#accept alert
a.accept()
#driver quit
driver.quit()"""

from time import time
from selenium import webdriver
from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.common.by import By


def test_pauses(driver):
    driver.get('https://selenium.dev/selenium/web/mouse_interaction.html')

    start = time()

    clickable = driver.find_element(By.ID, "clickable")
    ActionChains(driver) \
        .move_to_element(clickable) \
        .pause(1) \
        .click_and_hold() \
        .pause(1) \
        .send_keys("abc") \
        .perform()

    duration = time() - start
    assert duration > 2
    assert duration < 3


def test_releases_all(driver):
    driver.get('https://selenium.dev/selenium/web/mouse_interaction.html')

    clickable = driver.find_element(By.ID, "clickable")
    ActionChains(driver) \
        .click_and_hold(clickable) \
        .key_down(Keys.SHIFT) \
        .key_down("a") \
        .perform()

    ActionBuilder(driver).clear_actions()

    ActionChains(driver).key_down('a').perform()

    assert clickable.get_attribute('value')[0] == "A"
    assert clickable.get_attribute('value')[1] == "a"
