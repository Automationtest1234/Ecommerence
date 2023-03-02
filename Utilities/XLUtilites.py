import openpyxl
def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    workbook.active
    sheet=workbook[sheetName]
    return openpyxl.sheet.max_row

def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.max_column

def readData(file,sheetName,rownum,columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum,column=columno).value

def writeData(file,sheetName,rownum,columno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columno).value = data
    workbook.save(file)